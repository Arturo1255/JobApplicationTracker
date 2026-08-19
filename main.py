import os
from dataclasses import dataclass
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from email.utils import parsedate_to_datetime
from groq import Groq
import imaplib
import email
import json

llm_api_key = os.environ.get("GROQ_API_KEY")
client = Groq(llm_api_key)


@dataclass
class Application:
    company: str
    role: str
    status: str
    date_received: date
    email_subject: str
    last_updated: date


sample_emails = [
    {
        "subject": "Thank you for applying to Acme Corp!",
        "body": "We received your application for the Software Engineer position...",
        "date": "2026-07-15",
    },
    {
        "subject": "Interview invitation - Acme Corp",
        "body": "We'd like to schedule an interview for the Software Engineer role...",
        "date": "2026-07-22",
    },
    {
        "subject": "Update on your application to Globex",
        "body": "Unfortunately, we have decided to move forward with other candidates...",
        "date": "2026-07-18",
    },
]


def connect_to_mailbox(email_address: str, app_password: str, label: str = "Job_Applications"):
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(email_address, app_password)
    mail.select(f'"{label}"')
    return mail


def get_message_ids(mail) -> list:
    msgs = []
    status, message_ids = mail.search(None, "ALL")

    if status == "OK":
        msgs = message_ids[0].split()

    return msgs


def get_body(msg) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain":
                return part.get_payload(decode=True).decode(errors="ignore")
        return ""
    else:
        return msg.get_payload(decode=True).decode(errors="ignore")


def fetch_email(mail, msg_id) -> dict:
    fetched_email = {}
    status, msg_data = mail.fetch(msg_id, "(RFC822)")

    if status == "OK":
        raw_email = msg_data[0][1]
        parsed_email = email.message_from_bytes(raw_email)
        email_body = get_body(parsed_email)

        dt = parsedate_to_datetime(parsed_email["date"])
        date_string = dt.strftime("%Y-%m-%d")

        fetched_email.update({"subject": parsed_email["subject"], "body": email_body, "date": date_string})

        return fetched_email

    return None


def extract_company(subject: str) -> str:
    delimiters = [" to ", " at ", " - "]
    for delim in delimiters:
        if delim in subject:
            company = subject.split(delim)[-1].strip("!").strip(" ")
            return company
    return ""


def extract_status(body: str) -> str:
    lowerBody = body.lower()
    keywords = [
        ("unfortunately", "Rejected"),
        ("other candidates", "Rejected"),
        ("offer", "Offer"),
        ("interview", "Interview"),
        ("applied", "Applied"),
        ("received", "Applied"),
    ]

    for key, value in keywords:
        if key in lowerBody:
            return value

    return "Unknown"


def extract_role(body: str) -> str:
    delimiters = [" for the ", "for a "]
    cutoff = ["position", "role"]
    for delim in delimiters:
        if delim in body:
            role = body.split(delim)[-1].strip("! ")
            break
        else:
            role = ""

    for cut in cutoff:
        if cut in role:
            role = role.split(cut)[0].strip(" ")
            return role

    return ""


def extract_role_and_status_llm(subject: str, body: str) -> tuple[str, str]:
    try:
        response = client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[
                {
                    "role": "user",
                    "content": "Given the following subject and body return the Status and the role for the email in "
                               "json. The format must be exactly: {\"status\": \"value\", \"role\": \"value\"}. Also "
                               "the status must only be a value in this list Rejected, Interview, Applied and Offer \n "
                               + "Subject: " + subject + "\nBody: " + body
                }
            ],
            max_tokens=1000
        )
        print(repr(response.choices[0].message.content))
        data = json.loads(response.choices[0].message.content)

        return data["role"], data["status"]
    except Exception as e:
        print("Error: ")
        print(e)
        return extract_role(body), extract_status(body)


def parse_email(email: dict) -> Application:
    company = extract_company(email["subject"])
    role, status = extract_role_and_status_llm(email["subject"], email["body"])
    email_date = datetime.strptime(email["date"], "%Y-%m-%d").date()

    return Application(
        company=company,
        role=role,
        status=status,
        date_received=email_date,
        email_subject=email["subject"],
        last_updated=email_date
    )


def build_applications(emails: list[dict]) -> dict[tuple, Application]:
    applications = {}

    for email in emails:
        parsed = parse_email(email)
        key = (parsed.company, parsed.role)
        if key in applications:
            if applications[key].date_received > parsed.date_received:
                applications[key].date_received = parsed.date_received
            if applications[key].last_updated < parsed.last_updated:
                applications[key].status = parsed.status
                applications[key].last_updated = parsed.last_updated
        else:
            applications[key] = parsed
    return applications


def write_to_excel(applications: dict[tuple, Application], filename: str = "applications.xlsx"):
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    yellow_fill = PatternFill(start_color="FFEE8C", end_color="FFEE8C", fill_type="solid")

    color_fill = {
        "Rejected": red_fill,
        "Offer": green_fill,
        "Interview": yellow_fill,
        "Applied": no_fill
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    ws.append(["Company", "Role", "Status", "Date Applied", "Last Updated"])

    current_row = 1

    for key, app in applications.items():
        row = [app.company, app.role, app.status, app.date_received, app.last_updated]
        ws.append(row)
        current_row = current_row + 1

        for i in range(1, 6):
            cell = ws.cell(current_row, i)
            cell.fill = color_fill.get(app.status, no_fill)

    wb.save(filename)
    print(filename + " saved in current directory.")


def fetch_all_job_emails(email_address: str, app_password: str, label: str = "Job_Applications") -> list[dict]:
    all_emails = []
    mail_box = connect_to_mailbox(email_address, app_password, label)

    email_ids = get_message_ids(mail_box)

    for email_id in email_ids:
        fetched_email = fetch_email(mail_box, email_id)

        if fetched_email is not None:
            all_emails.append(fetched_email)

    return all_emails


# Press the green button in the gutter to run the script.

if __name__ == "__main__":
    email_address = os.environ.get("JOB_EMAIL_ADDRESS")
    app_password = os.environ.get("JOB_EMAIL_APP_PASSWORD")
    if not email_address or not app_password:
        raise ValueError("Set JOB_EMAIL_ADDRESS and JOB_EMAIL_APP_PASSWORD environment variables before running.")

    all_emails = fetch_all_job_emails(email_address, app_password)
    email_applications = build_applications(all_emails)
    write_to_excel(email_applications)

